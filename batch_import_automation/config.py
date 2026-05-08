"""
批量导入配置管理模块
从Excel文件和价格JSON文件读取配置
"""

import glob
import json
import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


class BatchImportConfigManager:
    """批量导入配置管理器"""
    
    @staticmethod
    def _find_excel_file(excel_file: str) -> str:
        """查找Excel文件，支持模糊匹配含"批量导入"的xlsx文件"""
        if os.path.exists(excel_file):
            return excel_file
        
        search_dir = os.path.dirname(excel_file) or "."
        matches = glob.glob(os.path.join(search_dir, "*批量导入*.xlsx"))
        if not matches:
            return excel_file
        
        matches.sort(key=os.path.getmtime, reverse=True)
        print(f"未找到 {excel_file}，自动匹配到: {matches[0]}")
        return matches[0]

    @staticmethod
    def _discover_excel_paths(excel_file: str, auto_discover: bool) -> list:
        """
        解析要加载的 Excel 路径列表。
        auto_discover 为 True 时：在 excel_file 所在目录查找所有「*批量导入*.xlsx」，
        若找到 2 个及以上则按文件名排序后全部返回（否则行为与单文件一致）。
        """
        if not auto_discover:
            return [BatchImportConfigManager._find_excel_file(excel_file)]

        abs_excel = os.path.abspath(excel_file)
        search_dir = os.path.dirname(abs_excel) or os.getcwd()
        matches = glob.glob(os.path.join(search_dir, "*批量导入*.xlsx"))
        matches = [os.path.normpath(p) for p in matches if os.path.isfile(p)]

        if len(matches) >= 2:
            matches.sort(key=lambda p: os.path.basename(p).lower())
            print(f"同目录下检测到 {len(matches)} 个「批量导入」表格，将按顺序全部处理：")
            for i, p in enumerate(matches, 1):
                print(f"   {i}. {p}")
            return matches
        if len(matches) == 1:
            return [matches[0]]

        return [BatchImportConfigManager._find_excel_file(excel_file)]
    
    def __init__(self, config_file: str = "属性配置.json", 
                 excel_file: str = "批量导入.xlsx",
                 price_file: str = "价格.json",
                 excel_files=None,
                 auto_discover_batch_excels: bool = False):
        """
        初始化配置管理器
        :param config_file: JSON配置文件路径（保留用于类目和属性）
        :param excel_file: 单个 Excel 路径（excel_files 未指定时使用；支持模糊匹配含"批量导入"的文件）
        :param price_file: 价格JSON文件路径
        :param excel_files: 可选，多个 Excel 路径列表，按顺序处理；指定后忽略 excel_file 与自动发现
        :param auto_discover_batch_excels: 仅在未指定 excel_files 时生效：为 True 时同目录多份
            「*批量导入*.xlsx」会全部加入队列（按文件名排序）
        """
        self.config_file = config_file
        self.price_file = price_file
        if excel_files is not None:
            if not excel_files:
                raise ValueError("excel_files 不能为空列表")
            self.excel_file_paths = [self._find_excel_file(p) for p in excel_files]
        else:
            self.excel_file_paths = self._discover_excel_paths(
                excel_file, auto_discover_batch_excels
            )
        self.excel_file = self.excel_file_paths[0]
        self._config = None
        self._excel_data = None
        self._price_data = None

    def switch_to_excel(self, excel_path: str) -> bool:
        """
        切换到指定 Excel 并清空已缓存的表格行数据（下次 get_import_config 会重新读取该文件）。
        :param excel_path: 已解析存在的 .xlsx 路径（通常来自 excel_file_paths）
        """
        if not excel_path or not os.path.exists(excel_path):
            print(f"Excel文件不存在: {excel_path}")
            return False
        self.excel_file = excel_path
        self._excel_data = None
        if hasattr(self, "_all_excel_data"):
            delattr(self, "_all_excel_data")
        return True
    
    def load_excel_data(self):
        """从Excel文件加载数据"""
        if not os.path.exists(self.excel_file):
            print(f"Excel文件不存在: {self.excel_file}")
            return None
        
        try:
            workbook = load_workbook(self.excel_file, data_only=True)
            # 默认读取第一个工作表
            worksheet = workbook.active
            
            # 读取Excel数据
            excel_data = {
                "站点": None,
                "采集词": None,
                "热搜词": None,
                "类目": {
                    "一级": None,
                    "二级": None,
                    "三级": None,
                    "四级": None,
                    "五级": None,
                },
                "重量": None,
                "店铺": None,
                "尺寸图": None,
                "PDF文件": None,
                "FDApdf文件": None,
                "TISpdf文件": None,
                "PS/ICCpdf文件": None,
                "属性": {
                    "材质": None,
                    "尺寸(长x宽x高)": None,
                    "适用年龄": None,
                    "FDA化妆品注册编号": None,
                    "TIS标准编号": None,
                    "PS/ICC编号": None,
                }
            }
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # 第一行是标题行，第二行开始是数据行
            # 读取第一行作为标题
            header_map = {}
            for col_idx in range(1, min(50, max_col + 1)):  # 扩展至50列，确保PDF等靠后列能被读取
                header_cell = worksheet.cell(row=1, column=col_idx)
                header_value = str(header_cell.value).strip() if header_cell.value else ""
                
                # 匹配标题（使用elif会导致只匹配第一个符合条件的）
                if "站点" in header_value and "站点" not in header_map:
                    header_map["站点"] = col_idx
                if "采集词" in header_value and "采集词" not in header_map:
                    header_map["采集词"] = col_idx
                if "热搜词" in header_value and "热搜词" not in header_map:
                    header_map["热搜词"] = col_idx
                if "重量" in header_value and "重量" not in header_map:
                    header_map["重量"] = col_idx
                if "一级" in header_value and "一级" not in header_map:
                    header_map["一级"] = col_idx
                if "二级" in header_value and "二级" not in header_map:
                    header_map["二级"] = col_idx
                if "三级" in header_value and "三级" not in header_map:
                    header_map["三级"] = col_idx
                if "四级" in header_value and "四级" not in header_map:
                    header_map["四级"] = col_idx
                if "五级" in header_value and "五级" not in header_map:
                    header_map["五级"] = col_idx
                if "店铺" in header_value and "店铺" not in header_map:
                    header_map["店铺"] = col_idx
                if "分组" in header_value and "分组" not in header_map:
                    header_map["分组"] = col_idx
                # 添加对属性的支持
                if "材质" in header_value and "材质" not in header_map:
                    header_map["材质"] = col_idx
                if "尺寸" in header_value and "长" in header_value and "宽" in header_value and "高" in header_value and "尺寸" not in header_map:
                    header_map["尺寸(长x宽x高)"] = col_idx
                if "适用年龄" in header_value and "适用年龄" not in header_map:
                    header_map["适用年龄"] = col_idx
                if "TIS标准编号" in header_value and "pdf" not in header_value.lower() and "TIS标准编号" not in header_map:
                    header_map["TIS标准编号"] = col_idx
                if (
                    "PS/ICC" in header_value
                    and "编号" in header_value
                    and "pdf" not in header_value.lower()
                    and "PS/ICC编号" not in header_map
                ):
                    header_map["PS/ICC编号"] = col_idx
                # 多列认证 PDF：必须在泛匹配「PDF文件」之前识别
                h_norm = header_value.replace(" ", "").lower()
                if "ps/icc" in header_value.lower() and "pdf" in header_value.lower() and "PS/ICCpdf文件" not in header_map:
                    header_map["PS/ICCpdf文件"] = col_idx
                elif ("tispdf" in h_norm or ("tis" in h_norm and "pdf" in h_norm and "标准" not in header_value)) and "TISpdf文件" not in header_map:
                    header_map["TISpdf文件"] = col_idx
                elif ("fdapdf" in h_norm or ("fda" in h_norm and "pdf" in h_norm and "化妆品注册" not in header_value)) and "FDApdf文件" not in header_map:
                    header_map["FDApdf文件"] = col_idx
                # PDF文件列：明确优先于FDA匹配。表格中的"PDF文件"列存的是PDF文件名（如"化妆品FDA.pdf"），
                # 不是FDA化妆品注册编号。含"PDF"或"认证文件"的列一律识别为PDF文件。
                if ("PDF文件" in header_value or "PDF 文件" in header_value) and "PDF文件" not in header_map:
                    header_map["PDF文件"] = col_idx
                elif ("PDF" in header_value or "认证文件" in header_value) and "PDF文件" not in header_map:
                    header_map["PDF文件"] = col_idx  # "PDF"、"FDA认证文件"、"CPR认证文件"等
                elif "FDA" in header_value and "化妆品注册编号" in header_value and "FDA化妆品注册编号" not in header_map:
                    # 仅当标题同时含"FDA"和"化妆品注册编号"时才映射为FDA化妆品注册编号，避免误匹配
                    header_map["FDA化妆品注册编号"] = col_idx
                if "尺寸图" in header_value and "尺寸图" not in header_map:
                    header_map["尺寸图"] = col_idx
            
            # 读取所有数据行（从第二行开始）
            all_excel_data = []
            max_data_row = min(max_row, 100)  # 最多读取100行
            last_site = None  # 保存上一行的站点，用于处理合并单元格
            
            for data_row in range(2, max_data_row + 1):
                # 检查这一行是否有数据（至少有一个非空单元格）
                has_data = False
                for col_idx in header_map.values():
                    cell = worksheet.cell(row=data_row, column=col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        has_data = True
                        break
                
                if not has_data:
                    continue  # 跳过空行
                
                # 读取这一行的数据
                row_data = {
                    "站点": None,
                    "采集词": None,
                    "热搜词": None,
                    "类目": {
                        "一级": None,
                        "二级": None,
                        "三级": None,
                        "四级": None,
                        "五级": None,
                    },
                    "重量": None,
                    "店铺": None,
                    "分组": None,
                    "尺寸图": None,
                    "PDF文件": None,
                    "FDApdf文件": None,
                    "TISpdf文件": None,
                    "PS/ICCpdf文件": None,
                    "属性": {
                        "材质": None,
                        "尺寸(长x宽x高)": None,
                        "适用年龄": None,
                        "FDA化妆品注册编号": None,
                        "TIS标准编号": None,
                        "PS/ICC编号": None,
                    }
                }
                
                for field, col_idx in header_map.items():
                    cell = worksheet.cell(row=data_row, column=col_idx)
                    value = cell.value
                    if value is not None:
                        value_str = str(value).strip()
                        if field == "站点":
                            row_data["站点"] = value_str
                            last_site = value_str  # 更新上一行的站点
                        elif field == "采集词":
                            row_data["采集词"] = value_str
                        elif field == "热搜词":
                            row_data["热搜词"] = value_str
                        elif field == "重量":
                            # 重量可能是数字，需要转换为字符串
                            if isinstance(value, (int, float)):
                                row_data["重量"] = str(value)
                            else:
                                row_data["重量"] = value_str
                        elif field in ["一级", "二级", "三级", "四级", "五级"]:
                            row_data["类目"][field] = value_str
                        elif field == "店铺":
                            row_data["店铺"] = value_str
                        elif field == "分组":
                            row_data["分组"] = value_str
                        elif field == "尺寸图":
                            row_data["尺寸图"] = value_str
                        elif field == "PDF文件":
                            row_data["PDF文件"] = value_str
                        elif field == "FDApdf文件":
                            row_data["FDApdf文件"] = value_str
                        elif field == "TISpdf文件":
                            row_data["TISpdf文件"] = value_str
                        elif field == "PS/ICCpdf文件":
                            row_data["PS/ICCpdf文件"] = value_str
                        elif field in [
                            "材质",
                            "尺寸(长x宽x高)",
                            "适用年龄",
                            "FDA化妆品注册编号",
                            "TIS标准编号",
                            "PS/ICC编号",
                        ]:
                            row_data["属性"][field] = value_str
                
                # 如果站点为空（合并单元格），使用上一行的站点
                if row_data["站点"] is None and last_site:
                    row_data["站点"] = last_site
                
                all_excel_data.append(row_data)
            
            # 如果有多行数据，返回所有数据；如果只有一行，保持兼容性
            if len(all_excel_data) > 0:
                excel_data = all_excel_data[0]  # 默认返回第一行（保持兼容）
                self._excel_data = excel_data
                self._all_excel_data = all_excel_data  # 保存所有数据
            
            # 尝试从下拉框获取类目选项（如果单元格有数据验证）
            # 检查工作表中的数据验证
            cell_coordinate = None
            for field in ["一级", "二级", "三级", "四级", "五级"]:
                if field in header_map:
                    col_idx = header_map[field]
                    cell_coordinate = worksheet.cell(row=data_row, column=col_idx).coordinate
                    
                    # 检查是否有数据验证（下拉框）
                    for dv in worksheet.data_validations.dataValidation:
                        # 检查数据验证是否应用于当前单元格
                        if dv.ranges and cell_coordinate in [str(r) for r in dv.ranges]:
                            if dv.type == 'list':
                                # 获取下拉框的选项
                                formula = dv.formula1 if hasattr(dv, 'formula1') else None
                                if formula:
                                    # 如果是引用范围，尝试解析
                                    if '!' in formula:
                                        # 格式如：Sheet2!$A$1:$A$10
                                        try:
                                            sheet_name, range_str = formula.split('!')
                                            range_str = range_str.replace('$', '')
                                            ref_sheet = workbook[sheet_name.strip()]
                                            # 解析范围
                                            if ':' in range_str:
                                                start_cell, end_cell = range_str.split(':')
                                                start_col = start_cell.rstrip('0123456789')
                                                start_row = int(start_cell[len(start_col):])
                                                end_col = end_cell.rstrip('0123456789')
                                                end_row = int(end_cell[len(end_col):])
                                                
                                                # 读取范围内的值
                                                options = []
                                                for row in range(start_row, end_row + 1):
                                                    cell_val = ref_sheet[f"{start_col}{row}"].value
                                                    if cell_val:
                                                        options.append(str(cell_val).strip())
                                                
                                                if options and not excel_data["类目"][field]:
                                                    # 如果当前单元格没有值，使用第一个选项
                                                    excel_data["类目"][field] = options[0]
                                        except Exception as e:
                                            # 忽略解析错误，继续处理
                                            pass
                                    elif formula.startswith('"') and formula.endswith('"'):
                                        # 如果是直接的值列表，如："选项1,选项2,选项3"
                                        options = [opt.strip() for opt in formula.strip('"').split(',')]
                                        if options and not excel_data["类目"][field]:
                                            excel_data["类目"][field] = options[0]
            
            workbook.close()
            
            print(f"已从Excel文件加载数据，共 {len(all_excel_data)} 条")
            # 避免打印时的编码问题，只显示基本信息
            for i in range(len(all_excel_data)):
                print(f"   第{i+1}条数据已加载")
            
            return excel_data
            
        except Exception as e:
            error_msg = f"读取Excel文件失败: {e}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            return None
    
    def load_price_data(self):
        """从价格JSON文件加载数据"""
        if not os.path.exists(self.price_file):
            print(f"价格文件不存在: {self.price_file}")
            return None
        
        try:
            with open(self.price_file, "r", encoding="utf-8") as f:
                price_data = json.load(f)
            
            self._price_data = price_data
            print("已从价格文件加载数据")
            return price_data
            
        except Exception as e:
            print(f"读取价格文件失败: {e}")
            return None
    
    def get_price_info(self, site: str, weight: str):
        """
        根据站点和重量获取价格信息
        :param site: 站点名称
        :param weight: 重量（字符串，如"0.5"）
        :return: (基础加价, 运费, 仓储费) 或 (None, None, None)
        """
        if not self._price_data:
            self.load_price_data()
        
        if not self._price_data:
            return None, None, None
        
        try:
            price_config = self._price_data.get("价格", {})
            site_config = price_config.get(site, {})
            
            if not site_config:
                print(f"未找到站点 '{site}' 的价格配置")
                for key in price_config.keys():
                    if site in key or key in site:
                        site_config = price_config.get(key, {})
                        if site_config:
                            print(f"使用相似站点名称: '{key}'")
                            break
                
                if not site_config:
                    return None, None, None

            def _lookup_weight_formula(cfg: dict, w_float: float, w_str: str):
                """
                从站点配置中查找某个重量档位的公式，兼容多种键名：
                - "1" / "1.0" / "1.00"
                - 带 kg 后缀："1kg" / "1.0kg"
                - 去尾零的浮点字符串
                """
                raw = str(w_str).strip()
                candidates = []
                if raw:
                    candidates.append(raw)

                # 标准化：保留 1 位小数（与 Excel 常见格式一致）
                try:
                    candidates.append(f"{float(raw):.1f}")
                except Exception:
                    # raw 可能不是数字，忽略
                    pass

                # 如果是整数重量（如 1.0），也尝试 "1"
                if abs(w_float - round(w_float)) < 1e-9:
                    candidates.append(str(int(round(w_float))))

                # 去掉尾随 0（如 "1.0" -> "1"、"1.00" -> "1"）
                if "." in raw:
                    candidates.append(raw.rstrip("0").rstrip("."))

                # 去重（保持顺序）
                deduped = []
                for c in candidates:
                    c = str(c).strip()
                    if c and c not in deduped:
                        deduped.append(c)

                for c in deduped:
                    if c in cfg:
                        return cfg.get(c), c
                    ckg = f"{c}kg"
                    if ckg in cfg:
                        return cfg.get(ckg), ckg
                return None, None
            
            # 查找对应重量的价格配置
            # 确保重量格式一致（去除可能的空格和单位）
            weight_clean = str(weight).strip()
            
            # 如果重量包含单位（如"0.4kg"），去掉单位
            if weight_clean.lower().endswith('kg'):
                weight_clean = weight_clean[:-2].strip()
            
            try:
                weight_float = float(weight_clean)
            except ValueError:
                print(f"重量格式不正确: {weight_clean}")
                return None, None, None
            
            # 如果重量大于1.0，需要特殊处理
            if weight_float > 1.0:
                # 基础加价固定为200
                base_price = "200"
                
                # 运费计算逻辑：
                # 整数部分：按「几个 1.0」计算 → 整数部分 × 1.0 的运费
                # 小数部分：按配置中对应档位计算 → 查 0.1～0.9 的运费
                int_part = int(weight_float)
                decimal_part = round(weight_float - int_part, 1)  # 保留一位小数，避免浮点误差
                
                # 获取1.0的重量配置
                weight_1_0_formula, matched_key_1_0 = _lookup_weight_formula(site_config, 1.0, "1.0")
                if not weight_1_0_formula:
                    print(f"未找到站点 '{site}' 重量 '1.0' 的价格配置")
                    print(f"可用重量键: {list(site_config.keys())[:10]}...")
                    return None, None, None
                
                # 解析1.0的运费
                parts_1_0 = weight_1_0_formula.split('+')
                if len(parts_1_0) < 2:
                    print(f"1.0的重量配置格式不正确: {weight_1_0_formula}")
                    return None, None, None
                
                shipping_1_0 = int(parts_1_0[1].strip())
                warehouse_fee = parts_1_0[2].strip() if len(parts_1_0) >= 3 else "10"
                
                # 整数部分运费
                total_shipping = int_part * shipping_1_0
                log_parts = [f"{int_part}×1.0[{matched_key_1_0}]的运费{shipping_1_0}"]
                
                # 小数部分：查 0.1～0.9 档位
                if decimal_part > 0:
                    decimal_formula, matched_key_decimal = _lookup_weight_formula(
                        site_config, decimal_part, f"{decimal_part:.1f}"
                    )
                    if not decimal_formula:
                        print(f"未找到站点 '{site}' 重量 '{decimal_part:.1f}' 的价格配置（0.1～0.9档位）")
                        print(f"可用重量键: {list(site_config.keys())[:10]}...")
                        return None, None, None
                    parts_decimal = decimal_formula.split('+')
                    if len(parts_decimal) < 2:
                        print(f"小数部分重量配置格式不正确: {decimal_formula}")
                        return None, None, None
                    shipping_decimal = int(parts_decimal[1].strip())
                    total_shipping += shipping_decimal
                    log_parts.append(f"{decimal_part:.1f}[{matched_key_decimal}]的运费{shipping_decimal}")
                
                print(f"找到价格配置（重量>1.0）: 基础加价={base_price}, 运费={total_shipping} ({'+'.join(log_parts)}), 仓储费={warehouse_fee}")
                return base_price, str(total_shipping), warehouse_fee
            
            # 重量<=1.0的情况，正常查找
            # 尝试多种格式匹配
            price_formula, matched_key = _lookup_weight_formula(site_config, weight_float, weight_clean)
            
            if not price_formula:
                print(f"未找到站点 '{site}' 重量 '{weight_clean}' 的价格配置")
                print(f"可用重量键: {list(site_config.keys())[:10]}...")
                return None, None, None
            
            # 解析价格公式，格式如："170%+12+10"
            # 第一部分是基础加价（去掉%），第二部分是运费，第三部分是仓储费
            parts = price_formula.split('+')
            if len(parts) >= 2:
                base_price = parts[0].rstrip('%').strip()
                shipping = parts[1].strip()
                warehouse_fee = parts[2].strip() if len(parts) >= 3 else "10"
                
                if matched_key and matched_key != weight_clean:
                    print(f"找到价格配置: 基础加价={base_price}, 运费={shipping}, 仓储费={warehouse_fee} (匹配键: {matched_key})")
                else:
                    print(f"找到价格配置: 基础加价={base_price}, 运费={shipping}, 仓储费={warehouse_fee}")
                return base_price, shipping, warehouse_fee
            else:
                print(f"价格公式格式不正确: {price_formula}")
                return None, None, None
                
        except Exception as e:
            print(f"解析价格信息失败: {e}")
            return None, None, None
    
    def load(self):
        """加载完整配置（包括JSON配置文件）"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, "r", encoding="utf-8") as f:
                    self._config = json.load(f)
            else:
                self._config = {}
            
            return self._config
        except Exception as e:
            print(f"读取配置文件失败: {e}")
            self._config = {}
            return self._config
    
    def reload(self):
        """重新加载配置"""
        self._config = None
        self._excel_data = None
        self._price_data = None
        if hasattr(self, "_all_excel_data"):
            delattr(self, "_all_excel_data")
        return self.load()
    
    def get_import_config(self, row_index=0):
        """
        获取导入配置（从Excel和价格文件）
        :param row_index: 数据行索引（0表示第一行数据）
        :return: 导入配置字典
        """
        # 加载Excel数据
        if not self._excel_data:
            self.load_excel_data()
        
        if not self._excel_data:
            print("无法加载Excel数据")
            return {}
        
        # 获取所有数据
        all_data = getattr(self, '_all_excel_data', [self._excel_data])
        
        if row_index >= len(all_data):
            print(f"行索引 {row_index} 超出范围（共 {len(all_data)} 条数据）")
            return {}
        
        excel_data = all_data[row_index]
        site = excel_data.get("站点")
        weight = excel_data.get("重量")
        
        # 从价格文件获取基础加价、运费和仓储费
        base_price, shipping, warehouse_fee = self.get_price_info(site, weight) if site and weight else (None, None, None)
        
        # 构建导入配置
        import_config = {
            "站点": site,
            "分组": excel_data.get("分组"),
            "采集词": excel_data.get("采集词"),
            "热搜词": excel_data.get("热搜词"),
            "重量": weight,
            "店铺": excel_data.get("店铺"),
            "链接": "",
            "尺寸图": excel_data.get("尺寸图"),
            "PDF文件": excel_data.get("PDF文件"),
            "FDApdf文件": excel_data.get("FDApdf文件"),
            "TISpdf文件": excel_data.get("TISpdf文件"),
            "PS/ICCpdf文件": excel_data.get("PS/ICCpdf文件"),
            "类目": excel_data.get("类目", {}),
            "属性": excel_data.get("属性", {})
        }
        
        if base_price:
            import_config["基础加价"] = base_price
        if shipping:
            import_config["运费"] = shipping
        if warehouse_fee:
            import_config["仓储费"] = warehouse_fee
        
        return import_config
    
    def get_all_import_configs(self):
        """
        获取所有行的导入配置
        :return: 导入配置列表
        """
        # 加载Excel数据
        if not self._excel_data:
            self.load_excel_data()
        
        all_data = getattr(self, '_all_excel_data', [self._excel_data])
        all_configs = []
        
        for i, excel_data in enumerate(all_data):
            config = self.get_import_config(i)
            if config:
                # 添加类目配置到导入配置中
                category_config = self.get_category_config(i)
                if any(category_config.values()):
                    config["类目"] = category_config
                
                # 添加属性配置到导入配置中（如果还没有）
                if "属性" not in config or not config["属性"]:
                    # 从Excel数据中获取属性配置
                    excel_data = all_data[i]
                    if "属性" in excel_data and excel_data["属性"]:
                        config["属性"] = excel_data["属性"]
                
                all_configs.append(config)
        
        return all_configs
    
    def get_category_config(self, row_index=0):
        """
        获取类目配置（从Excel文件）
        :param row_index: 数据行索引（0表示第一行数据）
        :return: 类目配置字典
        """
        if not self._excel_data:
            self.load_excel_data()
        
        if not self._excel_data:
            return {}
        
        # 获取所有数据
        all_data = getattr(self, '_all_excel_data', [self._excel_data])
        
        if row_index >= len(all_data):
            return {}
        
        return all_data[row_index].get("类目", {})
    
    def get_attribute_config(self):
        """
        获取属性配置（从JSON配置文件）
        :return: 属性配置字典
        """
        if not self._config:
            self.load()
        
        return self._config.get("属性", {})

