"""
将所有类目文本.json转换为WPS表格的级联下拉菜单（使用嵌套IF函数，不使用INDIRECT）
"""

import json
import re
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation


def clean_category_name(name):
    """清理类目名称，移除括号内容"""
    return re.sub(r'\([^)]*\)', '', name).strip()


def build_nested_if_formula(formula_parts, cell_ref):
    """
    构建嵌套IF函数公式
    :param formula_parts: [(条件值, 公式), ...] 列表
    :param cell_ref: 单元格引用，如 "A2"
    :return: 嵌套IF公式字符串（不含等号）
    """
    if not formula_parts:
        return '""'
    
    # 从最后一个开始嵌套
    last_condition, last_formula = formula_parts[-1]
    # 移除公式开头的等号
    if last_formula.startswith('='):
        last_formula = last_formula[1:]
    
    if_formula = f'IF({cell_ref}="{last_condition}",{last_formula},"")'
    
    # 从倒数第二个开始向前嵌套
    for i in range(len(formula_parts) - 2, -1, -1):
        condition, formula = formula_parts[i]
        # 移除公式开头的等号
        if formula.startswith('='):
            formula = formula[1:]
        if_formula = f'IF({cell_ref}="{condition}",{formula},{if_formula})'
    
    # 返回的公式不包含等号，调用者会添加
    return f'IF({cell_ref}="","",{if_formula})'


def extract_all_categories(data):
    """
    递归提取所有层级的类目数据
    返回结构化的字典
    """
    result = {
        'level1': [],
        'level2': {},  # {一级类目: [二级类目列表]}
        'level3': {},  # {一级_二级: [三级类目列表]}
        'level4': {}   # {一级_二级_三级: [四级类目列表]}
    }
    
    if not isinstance(data, dict):
        return result
    
    # 遍历一级类目
    for level1_key, level1_value in data.items():
        level1_clean = clean_category_name(level1_key)
        result['level1'].append(level1_clean)
        
        if isinstance(level1_value, dict) and level1_value:
            result['level2'][level1_clean] = []
            
            # 遍历二级类目
            for level2_key, level2_value in level1_value.items():
                level2_clean = clean_category_name(level2_key)
                result['level2'][level1_clean].append(level2_clean)
                
                level2_path = f"{level1_clean}_{level2_clean}"
                
                if isinstance(level2_value, dict) and level2_value:
                    result['level3'][level2_path] = []
                    
                    # 遍历三级类目
                    for level3_key, level3_value in level2_value.items():
                        level3_clean = clean_category_name(level3_key)
                        result['level3'][level2_path].append(level3_clean)
                        
                        level3_path = f"{level2_path}_{level3_clean}"
                        
                        if isinstance(level3_value, dict) and level3_value:
                            result['level4'][level3_path] = []
                            
                            # 遍历四级类目
                            for level4_key in level3_value.keys():
                                level4_clean = clean_category_name(level4_key)
                                result['level4'][level3_path].append(level4_clean)
    
    return result


def create_cascading_dropdown_excel(json_file, output_file):
    """
    创建带级联下拉菜单的Excel文件（WPS兼容版本）
    """
    import os
    
    # 检查文件是否存在
    abs_json_file = os.path.abspath(json_file)
    if not os.path.exists(abs_json_file):
        raise FileNotFoundError(f"文件不存在: {abs_json_file}")
    
    # 读取JSON数据
    print(f"[读取] 正在读取: {json_file}")
    print(f"[路径] 文件路径: {abs_json_file}")
    file_size = os.path.getsize(abs_json_file)
    print(f"[大小] 文件大小: {file_size} 字节")
    
    if file_size == 0:
        raise ValueError(f"文件为空: {abs_json_file}")
    
    try:
        with open(abs_json_file, 'r', encoding='utf-8') as f:
            content = f.read()
            # 移除BOM标记（如果有）
            if content.startswith('\ufeff'):
                content = content[1:]
            # 移除末尾空白
            content = content.strip()
            if not content:
                raise ValueError("文件内容为空")
            data = json.loads(content)
    except json.JSONDecodeError as e:
        print(f"[错误] JSON解析失败: {e}")
        print(f"[调试] 文件前200个字符: {repr(content[:200])}")
        raise
    
    # 提取类目数据
    print("[解析] 正在解析类目结构...")
    categories = extract_all_categories(data)
    
    print(f"   一级类目: {len(categories['level1'])} 个")
    print(f"   二级类目组: {len(categories['level2'])} 个")
    print(f"   三级类目组: {len(categories['level3'])} 个")
    print(f"   四级类目组: {len(categories['level4'])} 个")
    
    # 创建Excel工作簿
    wb = Workbook()
    
    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 创建主工作表（用于输入数据）
    ws_main = wb.create_sheet("数据输入", 0)
    ws_main['A1'] = "一级类目"
    ws_main['B1'] = "二级类目"
    ws_main['C1'] = "三级类目"
    ws_main['D1'] = "四级类目"
    
    # 设置表头样式
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col in range(1, 5):
        cell = ws_main.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 调整列宽
    ws_main.column_dimensions['A'].width = 25
    ws_main.column_dimensions['B'].width = 35
    ws_main.column_dimensions['C'].width = 35
    ws_main.column_dimensions['D'].width = 35
    
    # 创建辅助工作表（存储各级类目数据）
    ws_data = wb.create_sheet("类目数据", 1)
    
    # 设置A2单元格默认值为第一个一级类目（在提取数据后设置）
    
    # ========== 写入一级类目 ==========
    print("[创建] 正在创建一级类目...")
    ws_data['A1'] = "一级类目"
    for idx, cat in enumerate(categories['level1'], start=2):
        ws_data[f'A{idx}'] = cat
    
    level1_count = len(categories['level1'])
    # openpyxl的formula1需要包含等号，工作表名称用单引号包裹（WPS兼容）
    level1_formula = f"='类目数据'!$A$2:$A${level1_count + 1}"
    
    # 验证数据写入
    print(f"   [验证] 已写入一级类目到A2-A{level1_count + 1}")
    print(f"   [验证] 前5个一级类目: {categories['level1'][:5]}")
    
    # 设置A2单元格默认值为第一个一级类目
    if categories['level1']:
        ws_main['A2'] = categories['level1'][0]
        print(f"   [OK] 一级类目 ({level1_count} 项)，A2默认值: {categories['level1'][0]}")
        print(f"   [调试] 一级类目公式: {level1_formula}")
    else:
        print(f"   [OK] 一级类目 ({level1_count} 项)")
    
    # ========== 写入二级类目（每个一级类目一列，并创建命名区域） ==========
    print("[创建] 正在创建二级类目...")
    col = 2  # 从B列开始
    level2_formulas = {}  # {一级类目: 命名区域名称}
    level2_named_ranges = {}  # {一级类目: 命名区域名称}
    
    for level1 in categories['level1']:
        level2_list = categories['level2'].get(level1, [])
        if level2_list:
            # 写入二级类目数据
            ws_data.cell(1, col, f"{level1}_二级")
            for idx, level2 in enumerate(level2_list, start=2):
                ws_data.cell(idx, col, level2)
            
            # 创建命名区域（使用简短的名称）
            end_row = len(level2_list) + 1
            col_letter = chr(64 + col)  # A=65, B=66, ...
            range_formula = f"'类目数据'!${col_letter}$2:${col_letter}${end_row}"
            # 使用简短的命名区域名称（L2_列号）
            range_name = f"L2_{col}"
            try:
                wb.create_named_range(range_name, ws_data, range_formula)
                level2_formulas[level1] = range_name  # 存储命名区域名称
                level2_named_ranges[level1] = range_name
                print(f"   [OK] {level1} 的二级类目: {len(level2_list)} 项 (命名区域: {range_name})")
            except Exception as e:
                # 如果创建命名区域失败，回退到直接引用
                formula = range_formula
                level2_formulas[level1] = formula
                print(f"   [警告] 创建命名区域失败，使用直接引用: {level1} 的二级类目: {len(level2_list)} 项 (列{col_letter})")
            col += 1
    
    # ========== 写入三级类目（每个一级+二级组合一列，并创建命名区域） ==========
    print("[创建] 正在创建三级类目...")
    level3_formulas = {}  # {(一级, 二级): 命名区域名称}
    level3_named_ranges = {}  # {(一级, 二级): 命名区域名称}
    
    for level1 in categories['level1']:
        level2_list = categories['level2'].get(level1, [])
        for level2 in level2_list:
            level2_path = f"{level1}_{level2}"
            level3_list = categories['level3'].get(level2_path, [])
            
            if level3_list:
                # 写入三级类目数据
                ws_data.cell(1, col, f"{level1}_{level2}_三级")
                for idx, level3 in enumerate(level3_list, start=2):
                    ws_data.cell(idx, col, level3)
                
                # 创建命名区域（使用简短的名称）
                end_row = len(level3_list) + 1
                col_letter = chr(64 + col)
                range_formula = f"'类目数据'!${col_letter}$2:${col_letter}${end_row}"
                # 使用简短的命名区域名称（L3_列号）
                range_name = f"L3_{col}"
                try:
                    wb.create_named_range(range_name, ws_data, range_formula)
                    level3_formulas[(level1, level2)] = range_name  # 存储命名区域名称
                    level3_named_ranges[(level1, level2)] = range_name
                    print(f"   [OK] {level1}_{level2} 的三级类目: {len(level3_list)} 项 (命名区域: {range_name})")
                except Exception as e:
                    # 如果创建命名区域失败，回退到直接引用
                    formula = range_formula
                    level3_formulas[(level1, level2)] = formula
                    print(f"   [警告] 创建命名区域失败，使用直接引用: {level1}_{level2} 的三级类目: {len(level3_list)} 项 (列{col_letter})")
                col += 1
    
    # ========== 写入四级类目（每个一级+二级+三级组合一列，并创建命名区域） ==========
    print("[创建] 正在创建四级类目...")
    level4_formulas = {}  # {(一级, 二级, 三级): 命名区域名称}
    level4_named_ranges = {}  # {(一级, 二级, 三级): 命名区域名称}
    
    for level1 in categories['level1']:
        level2_list = categories['level2'].get(level1, [])
        for level2 in level2_list:
            level2_path = f"{level1}_{level2}"
            level3_list = categories['level3'].get(level2_path, [])
            
            for level3 in level3_list:
                level3_path = f"{level2_path}_{level3}"
                level4_list = categories['level4'].get(level3_path, [])
                
                if level4_list:
                    # 写入四级类目数据
                    start_row = 2
                    ws_data.cell(1, col, f"{level1}_{level2}_{level3}_四级")
                    for idx, level4 in enumerate(level4_list, start=2):
                        ws_data.cell(idx, col, level4)
                    
                    # 创建命名区域（使用简短的名称）
                    end_row = len(level4_list) + 1
                    col_letter = chr(64 + col)
                    range_formula = f"'类目数据'!${col_letter}${start_row}:${col_letter}${end_row}"
                    # 使用简短的命名区域名称（L4_列号）
                    range_name = f"L4_{col}"
                    try:
                        wb.create_named_range(range_name, ws_data, range_formula)
                        level4_formulas[(level1, level2, level3)] = range_name  # 存储命名区域名称
                        level4_named_ranges[(level1, level2, level3)] = range_name
                        print(f"   [OK] {level1}_{level2}_{level3} 的四级类目: {len(level4_list)} 项 (命名区域: {range_name})")
                    except Exception as e:
                        # 如果创建命名区域失败，回退到直接引用
                        formula = range_formula
                        level4_formulas[(level1, level2, level3)] = (col, formula)
                        print(f"   [警告] 创建命名区域失败，使用直接引用: {level1}_{level2}_{level3} 的四级类目: {len(level4_list)} 项 (列{col_letter})")
                    col += 1
    
    # ========== 设置数据验证（级联下拉菜单） ==========
    print("[设置] 正在设置数据验证...")
    
    # 一级类目：直接引用工作表范围
    # formula1应该包含等号
    dv_level1 = DataValidation(type="list", formula1=level1_formula)
    dv_level1.error = "请从下拉列表中选择一级类目"
    dv_level1.errorTitle = "输入错误"
    dv_level1.prompt = "请选择一级类目"
    dv_level1.promptTitle = "选择类目"
    dv_level1.showDropDown = True
    ws_main.add_data_validation(dv_level1)
    dv_level1.add("A2:A10000")
    print(f"   [OK] 一级类目数据验证已设置")
    
    # 二级类目：使用嵌套IF函数引用命名区域（简化公式长度）
    # 确保所有一级类目都在公式中，即使没有二级类目也要处理
    level2_formula_parts = []
    for level1 in categories['level1']:
        if level1 in level2_formulas:
            # 有二级类目，使用命名区域名称
            range_name = level2_formulas[level1]
            level2_formula_parts.append((level1, range_name))
        else:
            # 没有二级类目，使用空范围
            level2_formula_parts.append((level1, "'类目数据'!$A$2:$A$2"))  # 空范围
    
    # 使用命名区域名称构建公式（更短）
    level2_formula = build_nested_if_formula(level2_formula_parts, "A2")
    
    print(f"   [调试] 二级类目公式: ={level2_formula}")
    print(f"   [调试] 二级类目公式长度: {len(level2_formula)} 字符")
    
    # 如果公式仍然超过200字符，直接使用命名区域名称（不带INDIRECT，WPS可能不支持）
    if len(level2_formula) > 200:
        print(f"   [警告] 二级类目公式过长({len(level2_formula)}字符)，直接使用命名区域名称")
        # 重新构建，直接使用命名区域名称（不带INDIRECT）
        level2_formula_parts_direct = []
        for level1 in categories['level1']:
            if level1 in level2_named_ranges:
                range_name = level2_named_ranges[level1]
                level2_formula_parts_direct.append((level1, range_name))  # 直接使用命名区域名称
            else:
                level2_formula_parts_direct.append((level1, "'类目数据'!$A$2:$A$2"))
        level2_formula = build_nested_if_formula(level2_formula_parts_direct, "A2")
        print(f"   [调试] 使用命名区域后公式长度: {len(level2_formula)} 字符")
    
    # 检查公式是否为空或无效
    if not level2_formula or level2_formula == '""':
        print(f"   [警告] 二级类目公式为空，使用默认空范围")
        level2_formula = "'类目数据'!$A$2:$A$2"
    
    dv_level2 = DataValidation(type="list", formula1=f"={level2_formula}")
    dv_level2.error = "请先选择一级类目，然后从下拉列表中选择二级类目"
    dv_level2.errorTitle = "输入错误"
    dv_level2.prompt = "请先选择一级类目"
    dv_level2.promptTitle = "选择类目"
    dv_level2.showDropDown = True
    ws_main.add_data_validation(dv_level2)
    dv_level2.add("B2:B10000")
    print(f"   [OK] 二级类目数据验证已设置")
    
    # 三级类目：使用嵌套IF函数引用命名区域（简化公式长度）
    level3_formula_parts = []
    for (level1, level2), range_name in level3_formulas.items():
        # 构建条件：A2="一级" AND B2="二级"
        condition_key = (level1, level2)
        level3_formula_parts.append((condition_key, range_name))
    
    # 构建三级类目的嵌套IF公式
    if level3_formula_parts:
        # 从最后一个开始嵌套，直接使用命名区域名称（不带INDIRECT）
        last_key, last_range = level3_formula_parts[-1]
        level1_val, level2_val = last_key
        if_formula = f'IF(AND(A2="{level1_val}",B2="{level2_val}"),{last_range},"")'
        
        # 从倒数第二个开始向前嵌套
        for i in range(len(level3_formula_parts) - 2, -1, -1):
            key, range_name = level3_formula_parts[i]
            level1_val, level2_val = key
            if_formula = f'IF(AND(A2="{level1_val}",B2="{level2_val}"),{range_name},{if_formula})'
        
        level3_formula = f'IF(OR(A2="",B2=""),"",{if_formula})'
    else:
        level3_formula = '""'
    
    # openpyxl的formula1需要包含等号
    print(f"   [调试] 三级类目公式: ={level3_formula}")
    print(f"   [调试] 三级类目公式长度: {len(level3_formula)} 字符")
    
    # 如果公式仍然超过200字符，限制处理的组合数量
    if len(level3_formula) > 200:
        print(f"   [警告] 三级类目公式过长({len(level3_formula)}字符)，限制处理的组合数量")
        # 只处理前10个组合
        level3_formula_parts_limited = level3_formula_parts[:10]
        if level3_formula_parts_limited:
            last_key, last_range = level3_formula_parts_limited[-1]
            level1_val, level2_val = last_key
            if_formula = f'IF(AND(A2="{level1_val}",B2="{level2_val}"),INDIRECT("{last_range}"),"")'
            for i in range(len(level3_formula_parts_limited) - 2, -1, -1):
                key, range_name = level3_formula_parts_limited[i]
                level1_val, level2_val = key
                if_formula = f'IF(AND(A2="{level1_val}",B2="{level2_val}"),INDIRECT("{range_name}"),{if_formula})'
            level3_formula = f'IF(OR(A2="",B2=""),"",{if_formula})'
            print(f"   [调试] 限制后公式长度: {len(level3_formula)} 字符")
    
    # 检查公式是否为空或无效
    if not level3_formula or level3_formula == '""':
        print(f"   [警告] 三级类目公式为空，使用默认空范围")
        level3_formula = "'类目数据'!$A$2:$A$2"
    
    dv_level3 = DataValidation(type="list", formula1=f"={level3_formula}")
    dv_level3.error = "请先选择一级和二级类目，然后从下拉列表中选择三级类目"
    dv_level3.errorTitle = "输入错误"
    dv_level3.prompt = "请先选择一级和二级类目"
    dv_level3.promptTitle = "选择类目"
    dv_level3.showDropDown = True
    ws_main.add_data_validation(dv_level3)
    dv_level3.add("C2:C10000")
    print(f"   [OK] 三级类目数据验证已设置")
    
    # 四级类目：使用嵌套IF函数引用命名区域（简化公式长度）
    level4_formula_parts = []
    for (level1, level2, level3), range_name in level4_formulas.items():
        condition_key = (level1, level2, level3)
        level4_formula_parts.append((condition_key, range_name))
    
    # 构建四级类目的嵌套IF公式（限制组合数量以确保不超过255字符）
    # 只处理前10个组合，确保公式长度在255字符以内
    max_combinations = 10
    if len(level4_formula_parts) > max_combinations:
        print(f"   [警告] 四级类目组合过多({len(level4_formula_parts)}个)，只处理前{max_combinations}个")
        level4_formula_parts = level4_formula_parts[:max_combinations]
    
    if level4_formula_parts:
        # 从最后一个开始嵌套，直接使用命名区域名称（不带INDIRECT）
        last_key, last_range = level4_formula_parts[-1]
        level1_val, level2_val, level3_val = last_key
        if_formula = f'IF(AND(A2="{level1_val}",B2="{level2_val}",C2="{level3_val}"),{last_range},"")'
        
        # 从倒数第二个开始向前嵌套
        for i in range(len(level4_formula_parts) - 2, -1, -1):
            key, range_name = level4_formula_parts[i]
            level1_val, level2_val, level3_val = key
            if_formula = f'IF(AND(A2="{level1_val}",B2="{level2_val}",C2="{level3_val}"),{range_name},{if_formula})'
        
        level4_formula = f'IF(OR(A2="",B2="",C2=""),"",{if_formula})'
        
        # 检查公式长度
        formula_length = len(level4_formula)
        print(f"   [调试] 四级类目公式长度: {formula_length} 字符")
        
        # 如果仍然超过250字符，进一步限制
        if formula_length > 250:
            print(f"   [警告] 公式仍然过长，只保留前5个组合")
            level4_formula_parts = level4_formula_parts[:5]
            if level4_formula_parts:
                last_key, last_range = level4_formula_parts[-1]
                level1_val, level2_val, level3_val = last_key
                if_formula = f'IF(AND(A2="{level1_val}",B2="{level2_val}",C2="{level3_val}"),INDIRECT("{last_range}"),"")'
                for i in range(len(level4_formula_parts) - 2, -1, -1):
                    key, range_name = level4_formula_parts[i]
                    level1_val, level2_val, level3_val = key
                    if_formula = f'IF(AND(A2="{level1_val}",B2="{level2_val}",C2="{level3_val}"),INDIRECT("{range_name}"),{if_formula})'
                level4_formula = f'IF(OR(A2="",B2="",C2=""),"",{if_formula})'
                print(f"   [调试] 最终公式长度: {len(level4_formula)} 字符")
    else:
        level4_formula = '""'
    
    # openpyxl的formula1需要包含等号
    print(f"   [调试] 四级类目公式: ={level4_formula}")
    
    # 检查公式是否为空或无效
    if not level4_formula or level4_formula == '""':
        print(f"   [警告] 四级类目公式为空，使用默认空范围")
        level4_formula = "'类目数据'!$A$2:$A$2"
    
    dv_level4 = DataValidation(type="list", formula1=f"={level4_formula}")
    dv_level4.error = "请先选择一级、二级和三级类目，然后从下拉列表中选择四级类目"
    dv_level4.errorTitle = "输入错误"
    dv_level4.prompt = "请先选择一级、二级和三级类目"
    dv_level4.promptTitle = "选择类目"
    dv_level4.showDropDown = True
    ws_main.add_data_validation(dv_level4)
    dv_level4.add("D2:D10000")
    print(f"   [OK] 四级类目数据验证已设置")
    
    # ========== 设置自动填充默认值公式 ==========
    print("[设置] 正在设置自动填充默认值...")
    
    # B2: 根据A2自动填充二级类目的第一个
    if level2_formulas:
        # 构建公式：根据A2的值，返回对应二级类目的第一个
        b2_formula_parts = []
        for level1, range_name in level2_formulas.items():
            # 获取该一级类目的二级类目列表
            level2_list = categories['level2'].get(level1, [])
            if level2_list:
                # 使用INDEX和INDIRECT获取第一个二级类目
                if level1 in level2_named_ranges:
                    # 使用命名区域
                    b2_formula_parts.append((level1, f"INDEX(INDIRECT(\"{range_name}\"),1)"))
                else:
                    # 回退到直接引用
                    if isinstance(range_name, str) and '!' in range_name:
                        col_letter = range_name.split('!$')[1].split('$')[0]
                        b2_formula_parts.append((level1, f"INDEX('类目数据'!${col_letter}$2:${col_letter}$1000,1)"))
        
        if b2_formula_parts:
            # 构建嵌套IF公式
            last_level1, last_formula = b2_formula_parts[-1]
            b2_formula = f'IF(A2="{last_level1}",{last_formula},"")'
            for i in range(len(b2_formula_parts) - 2, -1, -1):
                level1, formula = b2_formula_parts[i]
                b2_formula = f'IF(A2="{level1}",{formula},{b2_formula})'
            b2_formula = f'=IF(A2="","",{b2_formula})'
            ws_main['B2'] = b2_formula
            print(f"   [OK] B2自动填充公式已设置")
    
    # C2: 根据A2和B2自动填充三级类目的第一个
    if level3_formulas:
        c2_formula_parts = []
        for (level1, level2), range_name in level3_formulas.items():
            level3_list = categories['level3'].get(f"{level1}_{level2}", [])
            if level3_list:
                if (level1, level2) in level3_named_ranges:
                    # 使用命名区域
                    c2_formula_parts.append(((level1, level2), f"INDEX(INDIRECT(\"{range_name}\"),1)"))
                else:
                    # 回退到直接引用
                    if isinstance(range_name, str) and '!' in range_name:
                        col_letter = range_name.split('!$')[1].split('$')[0]
                        c2_formula_parts.append(((level1, level2), f"INDEX('类目数据'!${col_letter}$2:${col_letter}$1000,1)"))
        
        if c2_formula_parts:
            last_key, last_formula = c2_formula_parts[-1]
            level1_val, level2_val = last_key
            c2_formula = f'IF(AND(A2="{level1_val}",B2="{level2_val}"),{last_formula},"")'
            for i in range(len(c2_formula_parts) - 2, -1, -1):
                key, formula = c2_formula_parts[i]
                level1_val, level2_val = key
                c2_formula = f'IF(AND(A2="{level1_val}",B2="{level2_val}"),{formula},{c2_formula})'
            c2_formula = f'=IF(OR(A2="",B2=""),"",{c2_formula})'
            ws_main['C2'] = c2_formula
            print(f"   [OK] C2自动填充公式已设置")
    
    # D2: 根据A2、B2和C2自动填充四级类目的第一个
    if level4_formulas:
        d2_formula_parts = []
        for (level1, level2, level3), range_name in level4_formulas.items():
            level3_path = f"{level1}_{level2}_{level3}"
            level4_list = categories['level4'].get(level3_path, [])
            if level4_list:
                if (level1, level2, level3) in level4_named_ranges:
                    # 使用命名区域
                    d2_formula_parts.append(((level1, level2, level3), f"INDEX(INDIRECT(\"{range_name}\"),1)"))
                else:
                    # 回退到直接引用
                    if isinstance(range_name, tuple):
                        col_num, formula = range_name
                        if isinstance(formula, str) and '!' in formula:
                            col_letter = formula.split('!$')[1].split('$')[0]
                            d2_formula_parts.append(((level1, level2, level3), f"INDEX('类目数据'!${col_letter}$2:${col_letter}$1000,1)"))
        
        if d2_formula_parts:
            # 只处理前10个组合（与数据验证保持一致）
            d2_formula_parts = d2_formula_parts[:10]
            last_key, last_formula = d2_formula_parts[-1]
            level1_val, level2_val, level3_val = last_key
            d2_formula = f'IF(AND(A2="{level1_val}",B2="{level2_val}",C2="{level3_val}"),{last_formula},"")'
            for i in range(len(d2_formula_parts) - 2, -1, -1):
                key, formula = d2_formula_parts[i]
                level1_val, level2_val, level3_val = key
                d2_formula = f'IF(AND(A2="{level1_val}",B2="{level2_val}",C2="{level3_val}"),{formula},{d2_formula})'
            d2_formula = f'=IF(OR(A2="",B2="",C2=""),"",{d2_formula})'
            ws_main['D2'] = d2_formula
            print(f"   [OK] D2自动填充公式已设置")
    
    # 将公式复制到其他行（第3行到第10000行）
    # 注意：使用相对引用，让公式自动适应行号
    b2_cell = ws_main['B2']
    c2_cell = ws_main['C2']
    d2_cell = ws_main['D2']
    
    b2_formula = b2_cell.value if b2_cell.value and isinstance(b2_cell.value, str) and b2_cell.value.startswith('=') else None
    c2_formula = c2_cell.value if c2_cell.value and isinstance(c2_cell.value, str) and c2_cell.value.startswith('=') else None
    d2_formula = d2_cell.value if d2_cell.value and isinstance(d2_cell.value, str) and d2_cell.value.startswith('=') else None
    
    if b2_formula or c2_formula or d2_formula:
        print(f"   [复制] 正在将公式复制到其他行...")
        for row in range(3, 10001):
            # B列公式：将A2替换为A{row}
            if b2_formula:
                b_formula = b2_formula.replace('A2', f'A{row}')
                ws_main[f'B{row}'] = b_formula
            
            # C列公式：将A2、B2替换为A{row}、B{row}
            if c2_formula:
                c_formula = c2_formula.replace('A2', f'A{row}').replace('B2', f'B{row}')
                ws_main[f'C{row}'] = c_formula
            
            # D列公式：将A2、B2、C2替换为A{row}、B{row}、C{row}
            if d2_formula:
                d_formula = d2_formula.replace('A2', f'A{row}').replace('B2', f'B{row}').replace('C2', f'C{row}')
                ws_main[f'D{row}'] = d_formula
        
        print(f"   [OK] 已将所有行的自动填充公式设置完成")
    
    # 保存文件
    print(f"\n[保存] 正在保存文件: {output_file}")
    try:
        wb.save(output_file)
        print(f"[完成] 文件已保存: {output_file}")
    except PermissionError:
        # 如果文件被占用，尝试使用临时文件名
        import time
        temp_file = f"{output_file}.temp_{int(time.time())}"
        print(f"[警告] 原文件被占用，保存到临时文件: {temp_file}")
        wb.save(temp_file)
        print(f"[完成] 临时文件已保存: {temp_file}")
        print(f"[提示] 请关闭原文件后，将临时文件重命名为: {output_file}")
    
    # 统计信息
    total_level2 = sum(len(v) for v in categories['level2'].values())
    total_level3 = sum(len(v) for v in categories['level3'].values())
    total_level4 = sum(len(v) for v in categories['level4'].values())
    
    print(f"\n[统计] 统计信息:")
    print(f"   一级类目: {len(categories['level1'])} 个")
    print(f"   二级类目: {total_level2} 个")
    print(f"   三级类目: {total_level3} 个")
    print(f"   四级类目: {total_level4} 个")


if __name__ == "__main__":
    import sys
    import io
    
    # 设置标准输出为UTF-8编码（Windows兼容）
    if sys.platform == 'win32':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    
    # 获取脚本所在目录
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)
    
    json_file = "文本.json"
    output_file = "类目级联下拉菜单.xlsx"
    
    if len(sys.argv) > 1:
        json_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    # 转换为绝对路径
    if not os.path.isabs(json_file):
        json_file = os.path.join(script_dir, json_file)
    if not os.path.isabs(output_file):
        output_file = os.path.join(script_dir, output_file)
    
    try:
        create_cascading_dropdown_excel(json_file, output_file)
        print("\n[完成] 完成！")
        print(f"[文件] 输出文件: {output_file}")
        print("\n[说明] 使用说明:")
        print("1. 打开生成的Excel文件（WPS表格或Excel）")
        print("2. 在'数据输入'工作表中，从A列开始选择类目")
        print("3. 选择一级类目后，二级类目会自动更新")
        print("4. 依次选择二级、三级、四级类目")
        print("5. '类目数据'工作表包含所有数据，可以隐藏但不要删除")
    except Exception as e:
        print(f"\n[错误] 错误: {e}")
        import traceback
        traceback.print_exc()

