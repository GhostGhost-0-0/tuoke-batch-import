# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import os

JSON_FILE = "文本.json"
OUTPUT_FILE = "WPS_五级级联下拉.xlsx"

if not os.path.exists(JSON_FILE):
    print(f"❌ 未找到 {JSON_FILE}")
    exit(1)

def safe_keys(obj):
    """安全获取字典的 keys，如果不是 dict 则返回空列表"""
    if isinstance(obj, dict):
        return list(obj.keys())
    return []

# 读取 JSON
with open(JSON_FILE, "r", encoding="utf-8") as f:
    raw_data = json.load(f)

# 构建映射（一至五级，结构与 extract_categories 生成的 文本.json 一致）
l1_to_l2 = {}
l2_to_l3 = {}
l3_to_l4 = {}
l4_to_l5 = {}

for l1, sub1 in raw_data.items():
    l2_list = safe_keys(sub1)
    l1_to_l2[l1] = l2_list
    for l2 in l2_list:
        sub2 = sub1[l2]
        l3_list = safe_keys(sub2)
        l2_to_l3[l2] = l3_list
        for l3 in l3_list:
            sub3 = sub2[l3]
            l4_list = safe_keys(sub3)
            l3_to_l4[l3] = l4_list
            for l4 in l4_list:
                sub4 = sub3[l4]
                l5_list = safe_keys(sub4)
                l4_to_l5[l4] = l5_list

# 获取所有唯一键（保持顺序）
l1_keys = list(raw_data.keys())
l2_keys = list(l2_to_l3.keys())
l3_keys = list(l3_to_l4.keys())
l4_keys = list(l4_to_l5.keys())

# 创建工作簿
wb = Workbook()
ws_main = wb.active
ws_main.title = "级联下拉"
ws_hidden = wb.create_sheet("数据源")
ws_hidden.sheet_state = "hidden"

# === 写入数据源表 ===

col = 1

# A列：一级分类（用于A列下拉）
ws_hidden.cell(1, col, "一级分类")
for i, key in enumerate(l1_keys, start=2):
    ws_hidden.cell(i, col, key)
col += 1

# B~...列：每个一级分类的二级选项
l2_start_col = col  # 记录二级分类列起始位置
for l1 in l1_keys:
    ws_hidden.cell(1, col, l1)
    for i, l2 in enumerate(l1_to_l2.get(l1, []), start=2):
        ws_hidden.cell(i, col, l2)
    col += 1
l2_end_col = col - 1  # 记录二级分类列结束位置

# 继续写入：每个二级分类的三级选项
l3_start_col = col  # 记录三级分类列起始位置
for l2 in l2_keys:
    ws_hidden.cell(1, col, l2)
    for i, l3 in enumerate(l2_to_l3.get(l2, []), start=2):
        ws_hidden.cell(i, col, l3)
    col += 1
l3_end_col = col - 1  # 记录三级分类列结束位置

# 继续写入：每个三级分类的四级选项
l4_start_col = col  # 记录四级分类列起始位置
for l3 in l3_keys:
    ws_hidden.cell(1, col, l3)
    for i, l4 in enumerate(l3_to_l4.get(l3, []), start=2):
        ws_hidden.cell(i, col, l4)
    col += 1
l4_end_col = col - 1  # 记录四级分类列结束位置

# 继续写入：每个四级分类的五级选项
l5_start_col = col  # 记录五级分类列起始位置
for l4 in l4_keys:
    ws_hidden.cell(1, col, l4)
    for i, l5 in enumerate(l4_to_l5.get(l4, []), start=2):
        ws_hidden.cell(i, col, l5)
    col += 1
l5_end_col = col - 1  # 记录五级分类列结束位置

# === 主表：设置级联下拉 ===
ws_main.append(["一级分类", "二级分类", "三级分类", "四级分类", "五级分类"])

# A列：一级
dv_a = DataValidation(type="list", formula1=f"数据源!$A$2:$A${len(l1_keys)+1}")
dv_a.add("A2:A1000")
ws_main.add_data_validation(dv_a)

# B列：依赖A列
formula_b = (
    "=OFFSET(数据源!$B$1,1,MATCH(A2,数据源!$1:$1,0)-2,"
    "COUNTA(OFFSET(数据源!$B$1,1,MATCH(A2,数据源!$1:$1,0)-2,100)),1)"
)
dv_b = DataValidation(type="list", formula1=formula_b)
dv_b.add("B2:B1000")
ws_main.add_data_validation(dv_b)

def col_num_to_letter(n):
    """将列号转换为列字母（1=A, 2=B, ..., 27=AA）"""
    result = ""
    while n > 0:
        n -= 1
        result = chr(65 + (n % 26)) + result
        n //= 26
    return result

l3_start_letter = col_num_to_letter(l3_start_col)
l3_end_letter = col_num_to_letter(l3_end_col)
l4_start_letter = col_num_to_letter(l4_start_col)
l4_end_letter = col_num_to_letter(l4_end_col)
l5_start_letter = col_num_to_letter(l5_start_col)
l5_end_letter = col_num_to_letter(l5_end_col)

# C列：根据B2查找对应的三级分类列
formula_c = (
    f"=IF(B2=\"\",\"\","
    f"OFFSET(数据源!${l3_start_letter}$1,1,MATCH(B2,数据源!${l3_start_letter}$1:${l3_end_letter}$1,0)-1,"
    f"COUNTA(OFFSET(数据源!${l3_start_letter}$1,1,MATCH(B2,数据源!${l3_start_letter}$1:${l3_end_letter}$1,0)-1,100)),1))"
)
dv_c = DataValidation(type="list", formula1=formula_c)
dv_c.add("C2:C1000")
ws_main.add_data_validation(dv_c)

# D列：依赖C列（四级分类）
formula_d = (
    f"=IF(C2=\"\",\"\","
    f"OFFSET(数据源!${l4_start_letter}$1,1,MATCH(C2,数据源!${l4_start_letter}$1:${l4_end_letter}$1,0)-1,"
    f"COUNTA(OFFSET(数据源!${l4_start_letter}$1,1,MATCH(C2,数据源!${l4_start_letter}$1:${l4_end_letter}$1,0)-1,100)),1))"
)
dv_d = DataValidation(type="list", formula1=formula_d)
dv_d.add("D2:D1000")
ws_main.add_data_validation(dv_d)

# E列：依赖D列（五级分类）
formula_e = (
    f"=IF(D2=\"\",\"\","
    f"OFFSET(数据源!${l5_start_letter}$1,1,MATCH(D2,数据源!${l5_start_letter}$1:${l5_end_letter}$1,0)-1,"
    f"COUNTA(OFFSET(数据源!${l5_start_letter}$1,1,MATCH(D2,数据源!${l5_start_letter}$1:${l5_end_letter}$1,0)-1,100)),1))"
)
dv_e = DataValidation(type="list", formula1=formula_e)
dv_e.add("E2:E1000")
ws_main.add_data_validation(dv_e)

# 保存
try:
    wb.save(OUTPUT_FILE)
    print(f"[完成] 已生成文件：{OUTPUT_FILE}")
except PermissionError:
    import time
    temp_file = f"{OUTPUT_FILE}.temp_{int(time.time())}"
    wb.save(temp_file)
    print(f"[警告] 原文件被占用，已保存到临时文件：{temp_file}")
    print(f"[提示] 请关闭原文件后，将临时文件重命名为：{OUTPUT_FILE}")

print("[提示] 用 WPS 打开测试：A->B->C->D->E 应全部联动！")
print(f"[信息] 列范围：二级({col_num_to_letter(l2_start_col)}-{col_num_to_letter(l2_end_col)}), "
      f"三级({l3_start_letter}-{l3_end_letter}), 四级({l4_start_letter}-{l4_end_letter}), "
      f"五级({l5_start_letter}-{l5_end_letter})")
